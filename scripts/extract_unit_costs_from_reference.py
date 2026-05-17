"""One-time helper: extract unit costs from Reference's bid spreadsheet.

Walks the DETAIL sheet of test_data/Park_Avenue_Elementary_School.xlsx,
collects every WS/R-prefixed line item, and prints Python `UnitCost(...)`
entries on stdout — ready to paste into bid_engine.pricing.

Each printed entry uses trade="bricklayer" with hours_per_unit derived from
Reference's UNIT LABOR dollars by dividing by the bricklayer wage total
($108.02). This makes `hours_per_unit * wage.total` reproduce Reference's exact
unit-labor figure regardless of which trade actually performed the work in
his bid — labor calibration is collapsed onto a single reference rate.

Diagnostics go to stderr via the logging module. The generated code goes to
stdout, so you can do:
    python scripts/extract_unit_costs_from_reference.py > /tmp/entries.py
"""

from __future__ import annotations

import logging
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
SPREADSHEET = PROJECT_ROOT / "test_data" / "Park_Avenue_Elementary_School.xlsx"

# Bricklayer total = $56.58 base + $51.44 supplements (Orange County, NY).
# See bid_engine.pricing.DEFAULT_WAGES.
BRICKLAYER_TOTAL = 108.02

CODE_RE = re.compile(r"^([A-Z]+\d+)\s*:\s*(.*)$")

logging.basicConfig(level=logging.INFO, format="%(message)s", stream=sys.stderr)
logger = logging.getLogger("extract")


def normalize_unit(u: str) -> str:
    u = u.strip().upper().replace(".", "")
    if u in ("LF", "FT", "LIN FT", "LINEAR FT"):
        return "LF"
    if u in ("SF", "SQ FT", "SQFT"):
        return "SF"
    if u in ("EA", "EACH"):
        return "EA"
    if u in ("LS", "LUMP SUM"):
        return "LS"
    return u


def main() -> int:
    if not SPREADSHEET.exists():
        logger.error("spreadsheet not found: %s", SPREADSHEET)
        return 1
    try:
        wb = openpyxl.load_workbook(SPREADSHEET, data_only=True)
    except Exception as exc:
        logger.error("failed to open %s: %s", SPREADSHEET, exc)
        return 1
    if "DETAIL" not in wb.sheetnames:
        logger.error("expected DETAIL sheet; found %s", wb.sheetnames)
        return 1
    ws = wb["DETAIL"]

    # (code, unit) -> list of (unit_labor, unit_material, description, row)
    by_key: dict[tuple[str, str], list[tuple[float, float, str, int]]] = defaultdict(list)
    rows_scanned = 0
    rows_skipped_unparseable = 0

    for r in range(28, ws.max_row + 1):
        item_num = ws.cell(r, 1).value
        try:
            int(item_num)
        except (TypeError, ValueError):
            continue
        rows_scanned += 1
        desc = ws.cell(r, 5).value
        unit = ws.cell(r, 7).value
        unit_labor = ws.cell(r, 10).value
        unit_material = ws.cell(r, 11).value
        if not desc or not unit:
            continue
        m = CODE_RE.match(str(desc).strip())
        if not m:
            continue
        code = m.group(1)
        description = m.group(2).strip()
        try:
            ul = float(unit_labor) if unit_labor not in (None, "") else 0.0
            um = float(unit_material) if unit_material not in (None, "") else 0.0
        except (TypeError, ValueError) as exc:
            logger.warning("row %d (%s): non-numeric labor/material (%s) — skipping",
                           r, code, exc)
            rows_skipped_unparseable += 1
            continue
        by_key[(code, normalize_unit(str(unit)))].append((ul, um, description, r))

    logger.info("scanned %d ITEM #-bearing rows, parsed %d (code,unit)/row entries",
                rows_scanned, sum(len(v) for v in by_key.values()))
    if rows_skipped_unparseable:
        logger.warning("%d rows skipped due to malformed labor/material",
                       rows_skipped_unparseable)

    # For each (code, unit) pick a canonical entry. If multiple rows for the
    # same key disagree on unit_labor or unit_material, log a warning.
    canonical: dict[tuple[str, str], tuple[float, float, str]] = {}
    for key in sorted(by_key):
        entries = by_key[key]
        unique_lab = {round(e[0], 4) for e in entries}
        unique_mat = {round(e[1], 4) for e in entries}
        if len(unique_lab) > 1:
            logger.warning("%s %s: unit_labor varies across %d rows %s — using first row %d",
                           *key, len(entries), sorted(unique_lab), entries[0][3])
        if len(unique_mat) > 1:
            logger.warning("%s %s: unit_material varies across %d rows %s — using first row %d",
                           *key, len(entries), sorted(unique_mat), entries[0][3])
        ul, um, description, _ = entries[0]
        canonical[key] = (ul, um, description)

    # Group by code for the by-code lookup table; if a code spans multiple units,
    # emit one line per (code, unit) and a comment flagging the multi-unit case.
    codes_with_multi_unit = {
        code for code in {k[0] for k in canonical}
        if sum(1 for k in canonical if k[0] == code) > 1
    }

    print("# === Generated by scripts/extract_unit_costs_from_reference.py ===")
    print('# trade="bricklayer" + hours_per_unit = reference_unit_labor_$ / 108.02')
    print("# reproduces Reference's unit-labor dollars exactly. Material is taken")
    print("# verbatim from the spreadsheet's UNIT MATERIAL column.")
    print(f"# Source: {SPREADSHEET.name}")
    print(f"# Codes with multiple unit variants: {sorted(codes_with_multi_unit) or 'none'}")
    print()

    for (code, unit), (ul, um, description) in sorted(canonical.items()):
        hours = ul / BRICKLAYER_TOTAL
        # Trim description to a sensible length for inline comment.
        d_short = description if len(description) <= 60 else description[:57] + "..."
        marker = "" if code not in codes_with_multi_unit else f"  ({unit})"
        key_repr = f'("{code}", "{unit}")' if code in codes_with_multi_unit else f'"{code}"'
        print(f'    {key_repr}: UnitCost("{code}", "bricklayer", '
              f'{hours:.6f}, {um:.4f}, description="{d_short}"),{marker}')

    print()
    print(f"# Total: {len(canonical)} (code, unit) entries "
          f"covering {len({k[0] for k in canonical})} unique codes.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
