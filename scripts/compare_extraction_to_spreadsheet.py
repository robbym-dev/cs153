"""Compare AI-extracted quantities to a reference bid spreadsheet.

Aggregates extracted items by (code, unit) and diffs against the same key in
the spreadsheet's DETAIL sheet. By default points at the Park Avenue
marked-up baseline (the original validation run). Pass --extractions DIR to
compare a different extraction set against the same spreadsheet — e.g. the
raw-plan extraction in tests/extractions/park_ave_original/.
"""

from __future__ import annotations

import argparse
import re
from collections import defaultdict
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent

# Default extraction set: the Park Avenue marked-up baseline.
DEFAULT_EXTRACTION_FILES = (
    PROJECT_ROOT / "tests" / "extractions" / "page2.txt",
    PROJECT_ROOT / "tests" / "extractions" / "page3.txt",
    PROJECT_ROOT / "tests" / "baseline_page5.txt",
    PROJECT_ROOT / "tests" / "extractions" / "page6.txt",
)
DEFAULT_SPREADSHEET = PROJECT_ROOT / "test_data" / "Park_Avenue_Elementary_School.xlsx"
TOLERANCE = 0.05

CODE_RE = re.compile(r"^([A-Z]+\d+)\s*:")


def normalize_unit(u: str) -> str:
    u = u.strip().upper().replace(".", "")
    if u in ("LF", "FT", "LIN FT", "LINEAR FT"):
        return "LF"
    if u in ("SF", "SQ FT", "SQFT"):
        return "SF"
    if u in ("EA", "EACH"):
        return "EA"
    return u


def load_extraction(paths) -> dict:
    totals: dict = defaultdict(float)
    rows = 0
    for path in paths:
        try:
            content = Path(path).read_text()
        except OSError as exc:
            print(f"  warning: could not read {path}: {exc}")
            continue
        for line in content.splitlines():
            line = line.strip()
            if not line:
                continue
            parts = line.split("\t")
            if len(parts) != 3:
                continue
            code, qty_s, unit = parts
            try:
                qty = float(qty_s)
            except ValueError:
                continue
            totals[(code.strip(), normalize_unit(unit))] += qty
            rows += 1
    print(f"  loaded {rows} rows across {len(list(paths))} file(s)")
    return dict(totals)


def load_spreadsheet(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["DETAIL"]
    totals: dict = defaultdict(float)
    rows = 0
    for r in range(28, ws.max_row + 1):
        item_num = ws.cell(r, 1).value
        try:
            int(item_num)
        except (TypeError, ValueError):
            continue
        desc = ws.cell(r, 5).value
        qty = ws.cell(r, 6).value
        unit = ws.cell(r, 7).value
        if not desc or qty is None or not unit:
            continue
        m = CODE_RE.match(str(desc).strip())
        if not m:
            continue
        try:
            qty_f = float(qty)
        except (TypeError, ValueError):
            continue
        totals[(m.group(1), normalize_unit(str(unit)))] += qty_f
        rows += 1
    print(f"  loaded {rows} prefixed line items from DETAIL sheet")
    return dict(totals)


def _resolve_paths(extractions_arg: str | None) -> tuple[Path, ...]:
    if not extractions_arg:
        return DEFAULT_EXTRACTION_FILES
    target = Path(extractions_arg)
    if not target.exists():
        raise FileNotFoundError(f"--extractions path does not exist: {target}")
    if target.is_dir():
        files = tuple(sorted(target.glob("page*.txt")))
        if not files:
            raise FileNotFoundError(f"no page*.txt files in {target}")
        return files
    return (target,)


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument(
        "--extractions",
        help="directory of page*.txt extraction files, or a single file "
             "(default: the marked-up Park Avenue baseline)",
    )
    p.add_argument(
        "--spreadsheet",
        default=str(DEFAULT_SPREADSHEET),
        help=f"path to the reference bid .xlsx (default: {DEFAULT_SPREADSHEET.name})",
    )
    p.add_argument("--label", default="extraction", help="label for the extraction side")
    args = p.parse_args(argv)

    paths = _resolve_paths(args.extractions)
    print(f"Loading {args.label}...")
    for path in paths:
        print(f"  {path}")
    extr = load_extraction(paths)
    print(f"Loading spreadsheet: {args.spreadsheet}")
    sheet = load_spreadsheet(Path(args.spreadsheet))

    extr_keys = set(extr.keys())
    sheet_keys = set(sheet.keys())
    common = extr_keys & sheet_keys

    matches, diffs = [], []
    for k in sorted(common):
        e_qty, s_qty = extr[k], sheet[k]
        delta = e_qty - s_qty
        (matches if abs(delta) < TOLERANCE else diffs).append((k, e_qty, s_qty, delta))

    only_sheet = sorted(sheet_keys - extr_keys)
    only_extr = sorted(extr_keys - sheet_keys)

    line = "-" * 78

    def header():
        print(f"  {'CODE':6} {'UNIT':4}   {'EXTRACTED':>10} {'SHEET':>10} {'DELTA':>10}")

    print(f"\n{line}\nMATCHES ({len(matches)}):")
    header()
    for (code, unit), e, s, d in matches:
        print(f"  {code:6} {unit:4}   {e:>10.2f} {s:>10.2f} {d:>+10.3f}")

    print(f"\n{line}\nQUANTITY DIFFERS ({len(diffs)}):")
    header()
    for (code, unit), e, s, d in diffs:
        print(f"  {code:6} {unit:4}   {e:>10.2f} {s:>10.2f} {d:>+10.3f}")

    print(f"\n{line}\nIN SPREADSHEET ONLY ({len(only_sheet)}):")
    print(f"  {'CODE':6} {'UNIT':4}   {'SHEET':>10}")
    for code, unit in only_sheet:
        print(f"  {code:6} {unit:4}   {sheet[(code, unit)]:>10.2f}")

    print(f"\n{line}\nIN EXTRACTION ONLY ({len(only_extr)}):")
    print(f"  {'CODE':6} {'UNIT':4}   {'EXTRACTED':>10}")
    for code, unit in only_extr:
        print(f"  {code:6} {unit:4}   {extr[(code, unit)]:>10.2f}")

    total_keys = len(extr_keys | sheet_keys)
    match_rate = (len(matches) / total_keys) if total_keys else 0.0
    extr_total = sum(extr.values())
    sheet_total = sum(sheet.values())

    print(f"\n{line}\nSUMMARY ({args.label})")
    print(
        f"  unique (code, unit) keys:  extracted={len(extr_keys)}  "
        f"spreadsheet={len(sheet_keys)}  union={total_keys}"
    )
    print(
        f"  match rate (|delta| < {TOLERANCE}): "
        f"{len(matches)}/{total_keys} = {match_rate:.1%}"
    )
    print(
        f"  qty differs: {len(diffs)}    sheet-only: {len(only_sheet)}    "
        f"extr-only: {len(only_extr)}"
    )
    print(
        f"  aggregate quantity sum:  extracted={extr_total:.1f}  "
        f"spreadsheet={sheet_total:.1f}  delta={extr_total - sheet_total:+.1f}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
