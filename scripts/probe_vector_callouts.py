"""Prototype: deterministic callout-circle detection + tiny-crop OCR.

Two auto-detection passes that generalize across engineering firms instead
of hardcoding the signature we observed on PA original page 8:

  1. Callout signature — collect every near-circular curve-only path on the
     page in a plausible callout-size range (10..60pt diameter), bin by
     (rounded diameter, rounded stroke width), pick the densest bin with
     a minimum-count threshold. That bin's (diameter, stroke) becomes the
     match signature within a small tolerance.

  2. Legend region — the keynote legend is a vertical column of regularly-
     spaced bullet circles. Bucket detected callouts by x-coordinate
     (50pt buckets) and score each bucket by (count × y-span × regularity
     of y-spacing). The best-scoring bucket above a minimum threshold is
     the legend column. Callouts in that column are excluded from the
     placed-callout count and reported separately.

Then for each remaining (placed) callout: render a 120x120pt crop at
600 DPI around its center via fitz.Page.get_pixmap(matrix=...) and ask
Claude (effort=low, no thinking, max_tokens=50) to read the code inside.
OCR in parallel via ThreadPoolExecutor.

Per-page OCR results cached as JSON; --reuse skips API calls.
"""

from __future__ import annotations

import argparse
import base64
import json
import logging
import os
import re
import sys
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import asdict, dataclass, field
from pathlib import Path

import fitz
import openpyxl
from anthropic import Anthropic

logger = logging.getLogger("probe_vector_callouts")

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_PDF = PROJECT_ROOT / "test_data" / "PA_Exterior_Renovations_Bid_Apr_6_2026.pdf"
DEFAULT_SPREADSHEET = PROJECT_ROOT / "test_data" / "Park_Avenue_Elementary_School.xlsx"
DEFAULT_OUTPUT_DIR = (
    PROJECT_ROOT / "tests" / "extractions" / "park_ave_original" / "vector_callouts"
)

# Search range for the callout signature detector. Diameters outside this band
# are either too small to carry a code label or too large to be keynote callouts.
MIN_DIAMETER = 10.0
MAX_DIAMETER = 60.0
ASPECT_RATIO_MAX = 1.2
MIN_SIGNATURE_CLUSTER = 8  # below this, no reliable signature

# Tolerance once a signature is picked
DIAMETER_TOL = 3.0   # ±pt
STROKE_TOL = 0.05    # ±pt

# Legend-column detector
LEGEND_BUCKET_WIDTH = 50.0   # pt
LEGEND_MIN_CIRCLES = 5
LEGEND_MIN_Y_SPAN_FRAC = 0.40  # of page height
LEGEND_GAP_TOL = 0.35          # how loose "regular spacing" can be

# Crop + OCR
CROP_HALF_PT = 60       # 120x120pt total
CROP_DPI = 600
OCR_MODEL = "claude-opus-4-7"
OCR_PROMPT = (
    "What keynote code is written inside this circle? Return ONLY the code, "
    "like WS5 or R2. Nothing else."
)
PARALLEL_WORKERS = 10
CODE_RE = re.compile(r"^([A-Z]+\d+[A-Z]*)$")


# ---------------------------------------------------------------------------
# Dataclasses
# ---------------------------------------------------------------------------


@dataclass
class Callout:
    cx: float
    cy: float
    diameter: float
    stroke: float
    in_legend: bool = False


@dataclass
class OCRResult:
    cx: float
    cy: float
    raw: str
    code: str | None
    in_legend: bool
    error: str | None = None


@dataclass
class PageResult:
    page: int
    signature: tuple[float, float] | None  # (diameter, stroke) or None
    legend_x_range: tuple[float, float] | None
    n_circles_total: int           # before legend filter
    n_placed: int                  # after legend filter
    n_legend: int                  # excluded as legend
    ocr_results: list[OCRResult] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Auto-detect callout signature
# ---------------------------------------------------------------------------


def collect_near_circles(page) -> list[Callout]:
    """Every near-circular curve-only path on the page within the plausible
    callout-size range. Doesn't filter by stroke/diameter — that's signature
    detection's job."""
    out: list[Callout] = []
    for p in page.get_drawings():
        rect = p["rect"]
        w = rect.x1 - rect.x0
        h = rect.y1 - rect.y0
        if min(w, h) <= 0:
            continue
        diameter = max(w, h)
        if not (MIN_DIAMETER <= diameter <= MAX_DIAMETER):
            continue
        if diameter / min(w, h) > ASPECT_RATIO_MAX:
            continue
        items = p.get("items", [])
        if not items or any(it[0] not in ("c", "h", "re") for it in items):
            continue
        if sum(1 for it in items if it[0] == "c") < 2:
            continue
        out.append(
            Callout(
                cx=(rect.x0 + rect.x1) / 2,
                cy=(rect.y0 + rect.y1) / 2,
                diameter=diameter,
                stroke=p.get("width") or 0.0,
            )
        )
    return out


def detect_signature(circles: list[Callout]) -> tuple[float, float] | None:
    """Bin by (round(diameter), round(stroke, 2)). Return the densest bin
    if its count clears MIN_SIGNATURE_CLUSTER, else None.
    """
    if not circles:
        return None
    bins: Counter = Counter()
    for c in circles:
        bins[(round(c.diameter), round(c.stroke, 2))] += 1
    (best_d, best_s), n = bins.most_common(1)[0]
    if n < MIN_SIGNATURE_CLUSTER:
        return None
    return float(best_d), float(best_s)


def filter_by_signature(
    circles: list[Callout], signature: tuple[float, float]
) -> list[Callout]:
    target_d, target_s = signature
    return [
        c for c in circles
        if abs(c.diameter - target_d) <= DIAMETER_TOL
        and abs(c.stroke - target_s) <= STROKE_TOL
    ]


# ---------------------------------------------------------------------------
# Auto-detect legend column
# ---------------------------------------------------------------------------


def detect_legend_column(
    callouts: list[Callout], page_height: float
) -> tuple[float, float] | None:
    """Return (x_min, x_max) of the legend column, or None.

    A legend column is a vertical run of callouts at near-constant x,
    spanning at least LEGEND_MIN_Y_SPAN_FRAC of page height, with regular
    vertical spacing.
    """
    if len(callouts) < LEGEND_MIN_CIRCLES:
        return None

    by_bucket: dict[float, list[Callout]] = defaultdict(list)
    for c in callouts:
        bucket = (c.cx // LEGEND_BUCKET_WIDTH) * LEGEND_BUCKET_WIDTH
        by_bucket[bucket].append(c)

    best_score = 0.0
    best_x_range: tuple[float, float] | None = None

    for bucket_x, items in by_bucket.items():
        if len(items) < LEGEND_MIN_CIRCLES:
            continue
        ys = sorted(c.cy for c in items)
        y_span = ys[-1] - ys[0]
        if y_span < page_height * LEGEND_MIN_Y_SPAN_FRAC:
            continue
        gaps = [ys[i + 1] - ys[i] for i in range(len(ys) - 1)]
        if not gaps:
            continue
        median_gap = sorted(gaps)[len(gaps) // 2]
        if median_gap <= 0:
            continue
        regularity = (
            sum(1 for g in gaps if abs(g - median_gap) < LEGEND_GAP_TOL * median_gap)
            / len(gaps)
        )
        # Score weights count and regularity equally.
        score = len(items) * regularity
        if score > best_score:
            best_score = score
            xs = sorted(c.cx for c in items)
            best_x_range = (xs[0] - 5, xs[-1] + 5)  # small padding
    return best_x_range


# ---------------------------------------------------------------------------
# OCR
# ---------------------------------------------------------------------------


def render_crop_png(page, callout: Callout) -> bytes:
    clip = (
        fitz.Rect(
            callout.cx - CROP_HALF_PT, callout.cy - CROP_HALF_PT,
            callout.cx + CROP_HALF_PT, callout.cy + CROP_HALF_PT,
        )
        & page.rect
    )
    if clip.is_empty or clip.is_infinite or clip.width < 1 or clip.height < 1:
        raise ValueError(
            f"degenerate crop rect for callout at ({callout.cx:.1f},{callout.cy:.1f})"
        )
    scale = CROP_DPI / 72.0
    mat = fitz.Matrix(scale, scale)
    pix = page.get_pixmap(matrix=mat, clip=clip, alpha=False)
    return pix.tobytes("png")


def ocr_one(client: Anthropic, png_bytes: bytes) -> str:
    b64 = base64.standard_b64encode(png_bytes).decode("ascii")
    msg = client.messages.create(
        model=OCR_MODEL,
        max_tokens=50,
        output_config={"effort": "low"},
        messages=[{
            "role": "user",
            "content": [
                {"type": "image",
                 "source": {"type": "base64", "media_type": "image/png", "data": b64}},
                {"type": "text", "text": OCR_PROMPT},
            ],
        }],
    )
    for block in msg.content:
        if block.type == "text":
            return block.text.strip()
    return ""


def _clean_code(raw: str) -> str | None:
    s = raw.strip().upper().strip("`'\".,:;")
    for token in reversed(re.split(r"[\s/,]+", s)):
        token = token.strip(".,;:()[]")
        m = CODE_RE.match(token)
        if m:
            return m.group(1)
    return None


def run_ocr_parallel(page, callouts: list[Callout]) -> list[OCRResult]:
    """OCR every callout (legend ones included — we log them but mark in_legend)."""
    client = Anthropic()
    crops: list[bytes | None] = []
    for c in callouts:
        try:
            crops.append(render_crop_png(page, c))
        except ValueError as exc:
            logger.warning("  skipping off-page callout: %s", exc)
            crops.append(None)

    results: list[OCRResult | None] = [None] * len(callouts)

    def worker(idx: int) -> tuple[int, OCRResult]:
        c = callouts[idx]
        if crops[idx] is None:
            return idx, OCRResult(c.cx, c.cy, "", None, c.in_legend, "degenerate crop")
        try:
            raw = ocr_one(client, crops[idx])
            return idx, OCRResult(c.cx, c.cy, raw, _clean_code(raw), c.in_legend)
        except Exception as exc:  # noqa: BLE001
            return idx, OCRResult(c.cx, c.cy, "", None, c.in_legend, str(exc))

    completed = 0
    with ThreadPoolExecutor(max_workers=PARALLEL_WORKERS) as ex:
        futures = [ex.submit(worker, i) for i in range(len(callouts))]
        for f in as_completed(futures):
            idx, res = f.result()
            results[idx] = res
            completed += 1
            if completed % 25 == 0 or completed == len(callouts):
                logger.info("    OCR progress: %d/%d", completed, len(callouts))
    return [r for r in results if r is not None]


# ---------------------------------------------------------------------------
# Per-page driver
# ---------------------------------------------------------------------------


def process_page(
    doc, page_number: int, reuse_path: Path | None
) -> PageResult:
    page = doc.load_page(page_number - 1)
    logger.info("page %d: %.0f x %.0f pts",
                page_number, page.rect.width, page.rect.height)

    circles = collect_near_circles(page)
    logger.info("  near-circular candidates: %d", len(circles))
    signature = detect_signature(circles)
    if signature is None:
        logger.warning("  no callout signature detected — skipping page")
        return PageResult(
            page=page_number, signature=None, legend_x_range=None,
            n_circles_total=0, n_placed=0, n_legend=0,
        )
    logger.info("  detected signature: diameter=%.1f pt, stroke=%.2f pt", *signature)
    callouts = filter_by_signature(circles, signature)
    logger.info("  signature-matched circles: %d", len(callouts))

    legend_x_range = detect_legend_column(callouts, page.rect.height)
    if legend_x_range:
        logger.info("  detected legend column: x ∈ (%.0f, %.0f)", *legend_x_range)
        for c in callouts:
            if legend_x_range[0] <= c.cx <= legend_x_range[1]:
                c.in_legend = True
    else:
        logger.info("  no legend column detected")
    n_legend = sum(1 for c in callouts if c.in_legend)
    n_placed = len(callouts) - n_legend
    logger.info("  placed callouts: %d   legend bullets: %d", n_placed, n_legend)

    if reuse_path and reuse_path.exists():
        logger.info("  reusing cached OCR results from %s", reuse_path)
        cached = json.loads(reuse_path.read_text())
        ocr_results = [OCRResult(**r) for r in cached["ocr_results"]]
    else:
        if not os.environ.get("ANTHROPIC_API_KEY"):
            raise RuntimeError("ANTHROPIC_API_KEY not set")
        ocr_results = run_ocr_parallel(page, callouts)
        if reuse_path:
            reuse_path.parent.mkdir(parents=True, exist_ok=True)
            payload = {
                "page": page_number,
                "signature": list(signature),
                "legend_x_range": list(legend_x_range) if legend_x_range else None,
                "n_circles_total": len(callouts),
                "n_placed": n_placed,
                "n_legend": n_legend,
                "ocr_results": [asdict(r) for r in ocr_results],
            }
            reuse_path.write_text(json.dumps(payload, indent=2))
            logger.info("  wrote %s", reuse_path)

    return PageResult(
        page=page_number,
        signature=signature,
        legend_x_range=legend_x_range,
        n_circles_total=len(callouts),
        n_placed=n_placed,
        n_legend=n_legend,
        ocr_results=ocr_results,
    )


# ---------------------------------------------------------------------------
# Tyler-spreadsheet comparison
# ---------------------------------------------------------------------------


SPREADSHEET_CODE_RE = re.compile(r"^([A-Z]+\d+)\s*:")


def load_tyler_totals(path: Path) -> dict[str, dict[str, float]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["DETAIL"]
    out: dict = defaultdict(lambda: defaultdict(float))
    for r in range(28, ws.max_row + 1):
        item = ws.cell(r, 1).value
        try:
            int(item)
        except (TypeError, ValueError):
            continue
        desc = ws.cell(r, 5).value
        qty = ws.cell(r, 6).value
        unit = ws.cell(r, 7).value
        if not desc or qty is None or not unit:
            continue
        m = SPREADSHEET_CODE_RE.match(str(desc).strip())
        if not m:
            continue
        try:
            qty_f = float(qty)
        except (TypeError, ValueError):
            continue
        u = str(unit).strip().upper().replace(".", "")
        if u in ("FT", "LIN FT", "LINEAR FT"):
            u = "LF"
        elif u in ("SQ FT", "SQFT"):
            u = "SF"
        out[m.group(1)][u] += qty_f
    return {k: dict(v) for k, v in out.items()}


def print_aggregated_summary(
    page_results: list[PageResult],
    tyler: dict[str, dict[str, float]],
):
    placed_counts: Counter = Counter()
    legend_counts: Counter = Counter()
    raw_outputs: Counter = Counter()
    by_page: dict[int, Counter] = {}
    n_ocr_total = 0
    n_ocr_valid = 0
    for pr in page_results:
        page_codes = Counter()
        for r in pr.ocr_results:
            raw_outputs[r.raw] += 1
            n_ocr_total += 1
            if r.code:
                n_ocr_valid += 1
                if r.in_legend:
                    legend_counts[r.code] += 1
                else:
                    placed_counts[r.code] += 1
                    page_codes[r.code] += 1
        by_page[pr.page] = page_codes

    print()
    print("=" * 95)
    print("MULTI-PAGE VECTOR-CALLOUT TAKEOFF vs TYLER")
    print("=" * 95)
    print()
    print(f"Pages processed: {[pr.page for pr in page_results]}")
    for pr in page_results:
        sig = (
            f"d={pr.signature[0]:.1f}pt s={pr.signature[1]:.2f}pt"
            if pr.signature else "—"
        )
        leg = (
            f"x∈({pr.legend_x_range[0]:.0f},{pr.legend_x_range[1]:.0f})"
            if pr.legend_x_range else "none"
        )
        print(
            f"  p{pr.page}: signature {sig}  legend {leg}  "
            f"placed={pr.n_placed}  legend_bullets={pr.n_legend}"
        )
    print(f"\nOCR valid / total: {n_ocr_valid}/{n_ocr_total} "
          f"({n_ocr_valid/n_ocr_total*100:.1f}%)" if n_ocr_total else "")

    print()
    print(f"{'CODE':6} {'PLACED':>6} {'LEGEND':>6}  per-page  Tyler quantities")
    print("-" * 95)
    all_codes = sorted(
        set(placed_counts) | set(legend_counts) | set(tyler),
        key=lambda c: (c[:1], int(re.sub(r"\D", "", c) or 0)),
    )
    ea_matched = 0
    ea_total = 0
    for code in all_codes:
        p = placed_counts.get(code, 0)
        leg = legend_counts.get(code, 0)
        per_page = ", ".join(
            f"p{pr.page}={by_page[pr.page].get(code, 0)}" for pr in page_results
        )
        tyler_row = tyler.get(code, {})
        tyler_summary = " + ".join(f"{q:g} {u}" for u, q in tyler_row.items()) or "—"
        marker = ""
        if "EA" in tyler_row and p > 0:
            ta = tyler_row["EA"]
            ea_total += 1
            delta = p - ta
            pct = (delta / ta * 100) if ta else 0
            if abs(pct) <= 15:
                marker = f"  Δ={delta:+.0f} ({pct:+.0f}%) ✓"
                ea_matched += 1
            else:
                marker = f"  Δ={delta:+.0f} ({pct:+.0f}%) ✗"
        print(f"{code:6} {p:>6} {leg:>6}   {per_page}   {tyler_summary}{marker}")

    print()
    if ea_total:
        print(f"EA-direct placed-callout comparisons within ±15%: {ea_matched}/{ea_total}")

    # Top unrecognized raw outputs
    bad = [(r, n) for r, n in raw_outputs.items() if not _clean_code(r)]
    bad.sort(key=lambda x: -x[1])
    if bad[:10]:
        print(f"\nTop unrecognized OCR outputs (top 10 of {len(bad)}):")
        for raw, n in bad[:10]:
            short = raw[:80].replace("\n", " ")
            print(f"  {n:3} × {short!r}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def _parse_pages(arg: str) -> list[int]:
    out = []
    for chunk in arg.split(","):
        chunk = chunk.strip()
        if not chunk:
            continue
        out.append(int(chunk))
    return out


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--pdf", default=str(DEFAULT_PDF))
    ap.add_argument("--pages", type=_parse_pages, default=[8],
                    help="comma-separated 1-indexed pages, e.g. 7,8,9,10")
    ap.add_argument("--spreadsheet", default=str(DEFAULT_SPREADSHEET))
    ap.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR))
    ap.add_argument("--suffix", default="_ocr_v2",
                    help="filename suffix for per-page OCR JSON")
    ap.add_argument("--reuse", action="store_true",
                    help="reuse cached OCR JSON if present (skip API calls)")
    args = ap.parse_args(argv)

    logging.basicConfig(level=logging.INFO, format="%(message)s")

    pdf_path = Path(args.pdf)
    if not pdf_path.exists():
        print(f"error: PDF not found: {pdf_path}", file=sys.stderr)
        return 1

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    doc = fitz.open(pdf_path)
    page_results = []
    for page_no in args.pages:
        reuse_path = output_dir / f"page{page_no}{args.suffix}.json"
        if not args.reuse and reuse_path.exists():
            reuse_path.unlink()  # force fresh run unless --reuse
        pr = process_page(doc, page_no, reuse_path)
        page_results.append(pr)
    doc.close()

    tyler = load_tyler_totals(Path(args.spreadsheet))
    print_aggregated_summary(page_results, tyler)
    return 0


if __name__ == "__main__":
    sys.exit(main())
