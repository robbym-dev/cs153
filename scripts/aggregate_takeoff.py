"""Aggregate multi-page takeoff JSON files and compare to a bid spreadsheet.

Reads tests/extractions/<project>/takeoff/page*_full.json (the output of
`python -m bid_engine.takeoff --full --json-out ...`) and produces a
combined-extraction vs spreadsheet comparison.

Aggregation rules:
  - Parsed quantities come from the keynote legend, which is reprinted on
    every elevation sheet → DEDUPE by (code, variant); warn if the same key
    differs across pages.
  - Probed quantities come from per-page callout counting → SUM by
    (code, unit) across pages.
  - When both parsed and probed exist for the same (code, base_bid) record,
    prefer the parsed value (the engineer's stated project total). Probed
    contribution is logged separately.
"""

from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_PAGES_DIR = PROJECT_ROOT / "tests" / "extractions" / "park_ave_original" / "takeoff"
DEFAULT_SPREADSHEET = PROJECT_ROOT / "test_data" / "Park_Avenue_Elementary_School.xlsx"
TOLERANCE_PCT = 15

CODE_RE = re.compile(r"^([A-Z]+\d+)\s*:")


def _normalize_unit(u: str) -> str:
    u = u.strip().upper().replace(".", "")
    if u in ("LF", "FT", "LIN FT", "LINEAR FT"):
        return "LF"
    if u in ("SF", "SQ FT", "SQFT"):
        return "SF"
    if u in ("EA", "EACH"):
        return "EA"
    return u


def load_tyler(path: Path) -> dict[tuple[str, str], float]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["DETAIL"]
    totals: dict = defaultdict(float)
    for r in range(28, ws.max_row + 1):
        item = ws.cell(r, 1).value
        try:
            int(item)
        except (TypeError, ValueError):
            continue
        desc = ws.cell(r, 5).value
        qty = ws.cell(r, 6).value
        unit = ws.cell(r, 7).value
        if not desc or qty is None or not unit:
            continue
        m = CODE_RE.match(str(desc).strip())
        if not m:
            continue
        try:
            qty_f = float(qty)
        except (TypeError, ValueError):
            continue
        totals[(m.group(1), _normalize_unit(str(unit)))] += qty_f
    return dict(totals)


def load_pages(pages_dir: Path, suffix: str = "_full") -> dict[int, dict]:
    pages = {}
    for f in sorted(pages_dir.glob(f"page*{suffix}.json")):
        data = json.loads(f.read_text())
        pages[data["page"]] = data
    return pages


def _apply_conversion(code: str, unit: str, qty: float, conversions: list[dict]) -> tuple[str, float, str]:
    """If a unit conversion exists for this (code, EA, qty), apply it.

    Conversions across pages should be consistent for the same code. If
    multiple pages produced a conversion for the same code, average their
    converted_quantity values. Returns (final_unit, final_qty, note).
    """
    if unit != "EA":
        return unit, qty, ""
    matching = [c for c in conversions if c["code"] == code]
    if not matching:
        return unit, qty, ""
    units = {c["correct_unit"] for c in matching}
    if len(units) > 1:
        # Pick the most common unit, average within that unit
        target = Counter(c["correct_unit"] for c in matching).most_common(1)[0][0]
    else:
        target = next(iter(units))
    same_unit = [c for c in matching if c["correct_unit"] == target]
    avg_qty = sum(c["converted_quantity"] for c in same_unit) / len(same_unit)
    return target, float(avg_qty), f"converted from {qty:g} EA via {len(same_unit)} page(s)"


def aggregate(pages: dict[int, dict]) -> tuple[dict, dict, list[str], list[dict]]:
    """Return (parsed_dedup, probed_sum, warnings, conversions).

    parsed_dedup keys: (code, variant) → {quantity, unit, pages: list[int]}
    probed_sum    keys: (code, unit)   → {quantity, callouts, pages: list[int]}
    conversions: list of unit-conversion records collected across pages
                 (each entry: {code, original_unit, original_quantity,
                 correct_unit, converted_quantity, reasoning, page})
    """
    parsed_dedup: dict = {}
    probed_sum: dict = defaultdict(lambda: {"quantity": 0.0, "callouts": 0, "pages": []})
    warnings: list[str] = []
    conversions: list[dict] = []

    for page, data in sorted(pages.items()):
        for code, pqs in data.get("parsed", {}).items():
            for pq in pqs:
                key = (code, pq["variant"])
                if key not in parsed_dedup:
                    parsed_dedup[key] = {
                        "quantity": pq["quantity"],
                        "unit": _normalize_unit(pq["unit"]),
                        "pages": [page],
                    }
                else:
                    existing = parsed_dedup[key]
                    existing["pages"].append(page)
                    if (
                        abs(existing["quantity"] - pq["quantity"]) > 0.01
                        or existing["unit"] != _normalize_unit(pq["unit"])
                    ):
                        warnings.append(
                            f"parsed {code} {pq['variant']} differs across pages: "
                            f"{existing['quantity']} {existing['unit']} (kept) "
                            f"vs {pq['quantity']} {pq['unit']} on page {page}"
                        )

        for code, probe in data.get("probed", {}).items():
            unit = _normalize_unit(probe["unit"])
            key = (code, unit)
            probed_sum[key]["quantity"] += float(probe["quantity"])
            probed_sum[key]["callouts"] += int(probe["callout_count"])
            probed_sum[key]["pages"].append(page)

        # v2 conversions (absent in v1 files)
        for code, conv in data.get("converted", {}).items():
            conversions.append({**conv, "page": page})

    return parsed_dedup, dict(probed_sum), warnings, conversions


def build_engine_totals(
    parsed_dedup: dict,
    probed_sum: dict,
    conversions: list[dict] | None = None,
) -> dict[tuple[str, str], dict]:
    """Combined engine extraction per (code, unit). Parsed (base_bid) wins
    where both sources exist; probed-only otherwise. Alternates kept separate.

    When `conversions` is provided (v2 output) and a code's combined unit is
    EA, replace with the converted unit/quantity from the unit-conversion pass.
    """
    engine: dict = {}
    conversions = conversions or []

    for (code, variant), v in parsed_dedup.items():
        if variant != "base_bid":
            continue
        unit, qty, note = _apply_conversion(code, v["unit"], v["quantity"], conversions)
        key = (code, unit)
        engine[key] = {
            "quantity": qty,
            "source": "parsed" + (f" [{note}]" if note else ""),
            "pages": v["pages"],
        }

    for (code, unit), v in probed_sum.items():
        new_unit, new_qty, note = _apply_conversion(code, unit, v["quantity"], conversions)
        key = (code, new_unit)
        if key in engine:
            engine[key]["source"] += " + probed"
        elif new_qty > 0 or v["callouts"] > 0:
            engine[key] = {
                "quantity": new_qty,
                "source": "probed" + (f" [{note}]" if note else ""),
                "pages": v["pages"],
            }
    return engine


def print_comparison(
    engine: dict,
    tyler: dict,
    parsed_dedup: dict,
    probed_sum: dict,
    warnings: list[str],
    pages: dict[int, dict],
):
    print("=" * 100)
    print("MULTI-PAGE TAKEOFF — PA ORIGINAL DRAWINGS vs TYLER")
    print("=" * 100)
    print()
    print(f"Pages aggregated: {sorted(pages.keys())}")
    print(
        f"Keynote records across pages: "
        + ", ".join(
            f"p{p}={len(d.get('keynotes', []))}" for p, d in sorted(pages.items())
        )
    )
    print(
        f"Per-page scale: "
        + ", ".join(
            f"p{p}={d['scale']['raw'] if isinstance(d['scale'], dict) else 'multi'}"
            for p, d in sorted(pages.items())
        )
    )
    if warnings:
        print(f"\nAggregation warnings ({len(warnings)}):")
        for w in warnings:
            print(f"  ! {w}")

    print(
        f"\nParsed records (deduped across {len(pages)} pages): {len(parsed_dedup)}"
    )
    print(f"Probed (code,unit) totals (summed across pages):   {len(probed_sum)}")

    # Line-by-line comparison vs Tyler
    all_keys = sorted(set(engine) | set(tyler))
    print()
    print(
        f"{'M':1} {'CODE':6} {'UNIT':4}  {'ENGINE':>9} {'TYLER':>9} "
        f"{'Δ':>9} {'Δ%':>7}  SOURCE"
    )
    print("-" * 100)
    matches_within_tolerance = 0
    comparable = 0
    extr_total_same_unit = 0.0
    tyler_total_same_unit = 0.0
    for key in all_keys:
        code, unit = key
        e_info = engine.get(key)
        e = e_info["quantity"] if e_info else 0.0
        src = e_info["source"] if e_info else "—"
        t = tyler.get(key, 0.0)
        if t > 0:
            comparable += 1
            extr_total_same_unit += e
            tyler_total_same_unit += t
            delta = e - t
            pct = delta / t * 100
            within = abs(pct) <= TOLERANCE_PCT
            if within:
                matches_within_tolerance += 1
            marker = "✓" if within else "✗"
            print(
                f"{marker} {code:6} {unit:4}  {e:>9.2f} {t:>9.2f} "
                f"{delta:>+9.2f} {pct:>+6.1f}%  {src}"
            )
        else:
            # Engine returned something Tyler doesn't have, or Tyler has a row
            # in a different unit
            print(
                f"  {code:6} {unit:4}  {e:>9.2f} {t:>9.2f} "
                f"{'—':>9} {'—':>7}  {src}"
            )

    print()
    print("=" * 100)
    print("SUMMARY")
    print("=" * 100)
    print(
        f"  Same-unit comparable codes:           {comparable}"
    )
    print(
        f"  Within ±{TOLERANCE_PCT}% of Tyler:                  "
        f"{matches_within_tolerance}/{comparable}"
        f" ({matches_within_tolerance/comparable*100:.0f}%)"
        if comparable
        else "  (no comparable rows)"
    )
    print(
        f"  Aggregate sum (same-unit comparable): "
        f"engine={extr_total_same_unit:.1f}  "
        f"tyler={tyler_total_same_unit:.1f}  "
        f"delta={extr_total_same_unit - tyler_total_same_unit:+.1f} "
        f"({(extr_total_same_unit-tyler_total_same_unit)/tyler_total_same_unit*100:+.1f}%)"
    )

    # Alternate-bid records
    alts = [(k, v) for k, v in parsed_dedup.items() if k[1] != "base_bid"]
    if alts:
        print()
        print("Parsed alternate-bid quantities (not summed into base totals):")
        for (code, variant), v in sorted(alts):
            print(
                f"  {code:6} {variant:14} {v['quantity']:>8.2f} {v['unit']:4} "
                f"(pages {v['pages']})"
            )


def main(argv=None) -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--pages-dir", default=str(DEFAULT_PAGES_DIR))
    p.add_argument("--spreadsheet", default=str(DEFAULT_SPREADSHEET))
    p.add_argument(
        "--suffix",
        default="_full",
        help="JSON filename suffix to load (_full for v1, _v2 for v2). Default: _full",
    )
    args = p.parse_args(argv)

    pages = load_pages(Path(args.pages_dir), suffix=args.suffix)
    if not pages:
        print(f"no page*{args.suffix}.json found in {args.pages_dir}")
        return 1
    parsed_dedup, probed_sum, warnings, conversions = aggregate(pages)
    tyler = load_tyler(Path(args.spreadsheet))
    engine = build_engine_totals(parsed_dedup, probed_sum, conversions)
    print_comparison(engine, tyler, parsed_dedup, probed_sum, warnings, pages)
    if conversions:
        print()
        print(f"Unit conversions applied ({len(conversions)} record(s) across pages):")
        seen = set()
        for c in conversions:
            sig = (c["code"], c["correct_unit"])
            if sig in seen:
                continue
            seen.add(sig)
            print(
                f"  {c['code']:6} {c['original_quantity']:>5.1f} {c['original_unit']} → "
                f"{c['converted_quantity']:>7.2f} {c['correct_unit']}  "
                f"{c['reasoning'][:70]}"
            )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
