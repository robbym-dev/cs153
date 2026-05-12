"""Tests for bid_engine.bid_generator.

The headline test loads the Park Avenue extraction data, runs it through
price_bid → generate_bid_excel → openpyxl read-back, and verifies the
TOTAL BASE BID cell on disk matches the in-memory Bid.total.
"""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path

import openpyxl
import pytest

from bid_engine.bid_generator import (
    COLUMN_SPECS,
    TOTAL_COST_COLUMN,
    BidHeader,
    generate_bid_excel,
)
from bid_engine.pricing import Bid, ScopeItem, price_bid

PROJECT_ROOT = Path(__file__).resolve().parent.parent
EXTRACTION_FILES = (
    PROJECT_ROOT / "tests" / "extractions" / "page2.txt",
    PROJECT_ROOT / "tests" / "extractions" / "page3.txt",
    PROJECT_ROOT / "tests" / "baseline_page5.txt",
    PROJECT_ROOT / "tests" / "extractions" / "page6.txt",
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _load_park_avenue_scope() -> list[ScopeItem]:
    """Reproduce the run_full_pipeline aggregation as a test fixture."""
    totals: dict[tuple[str, str], float] = defaultdict(float)
    for path in EXTRACTION_FILES:
        for line in path.read_text().splitlines():
            parts = line.strip().split("\t")
            if len(parts) != 3:
                continue
            code, qty, unit = parts
            totals[(code.strip(), unit.strip())] += float(qty)
    return [
        ScopeItem(code=code, quantity=qty, unit=unit)
        for (code, unit), qty in sorted(totals.items())
    ]


def _find_cell_by_value(ws, target: str):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == target:
                return cell
    return None


# ---------------------------------------------------------------------------
# Headline test: Park Avenue round-trip
# ---------------------------------------------------------------------------


def test_park_avenue_roundtrip_total_matches(tmp_path):
    scope = _load_park_avenue_scope()
    assert scope, "no scope items loaded from extraction files — check paths"
    bid = price_bid(scope)
    assert bid.line_items, "price_bid produced no line items — pricing module regressed"

    output = tmp_path / "park_ave_bid.xlsx"
    result = generate_bid_excel(
        bid,
        BidHeader(
            project_name="PARK AVENUE ELEMENTARY SCHOOL",
            address="10 Park Avenue, Warwick, NY 10990",
            date="2026-05-12",
        ),
        output,
    )
    assert result == output
    assert output.exists()

    wb = openpyxl.load_workbook(output, data_only=True)
    ws = wb.active

    # Header total
    header_total_cell = ws["B7"]
    assert header_total_cell.value == pytest.approx(bid.total)

    # TOTAL BASE BID row at the bottom
    label_cell = _find_cell_by_value(ws, "TOTAL BASE BID")
    assert label_cell is not None, "missing TOTAL BASE BID row"
    grand_total_cell = ws.cell(row=label_cell.row, column=TOTAL_COST_COLUMN)
    assert grand_total_cell.value == pytest.approx(bid.total)


# ---------------------------------------------------------------------------
# Structure / formatting
# ---------------------------------------------------------------------------


@pytest.fixture
def small_bid() -> Bid:
    return price_bid(
        [
            ScopeItem("WS5", 23.0, "EA"),
            ScopeItem("WS1", 100.0, "LF"),
        ]
    )


@pytest.fixture
def small_bid_xlsx(tmp_path, small_bid):
    out = tmp_path / "small.xlsx"
    generate_bid_excel(small_bid, BidHeader(project_name="Test Project"), out)
    return out, small_bid


def test_header_fields_populated(small_bid_xlsx):
    path, _bid = small_bid_xlsx
    ws = openpyxl.load_workbook(path, data_only=True).active
    assert ws["A1"].value == "GENERAL SUMMARY"
    assert ws["A3"].value == "PROJECT:"
    assert ws["B3"].value == "Test Project"
    assert ws["A7"].value == "TOTAL BASE BID:"


def test_column_headers_match_spec(small_bid_xlsx):
    path, _ = small_bid_xlsx
    ws = openpyxl.load_workbook(path).active
    for col_idx, (label, _width) in enumerate(COLUMN_SPECS, start=1):
        assert ws.cell(row=10, column=col_idx).value == label


def test_column_widths_set(small_bid_xlsx):
    path, _ = small_bid_xlsx
    ws = openpyxl.load_workbook(path).active
    from openpyxl.utils import get_column_letter

    for col_idx, (_label, expected_width) in enumerate(COLUMN_SPECS, start=1):
        actual = ws.column_dimensions[get_column_letter(col_idx)].width
        assert actual == expected_width


def test_column_headers_are_bold(small_bid_xlsx):
    path, _ = small_bid_xlsx
    ws = openpyxl.load_workbook(path).active
    for col_idx in range(1, len(COLUMN_SPECS) + 1):
        assert ws.cell(row=10, column=col_idx).font.bold is True


def test_line_items_written_in_order(small_bid_xlsx):
    path, bid = small_bid_xlsx
    ws = openpyxl.load_workbook(path, data_only=True).active
    for i, line in enumerate(bid.line_items, start=1):
        row = 10 + i
        assert ws.cell(row=row, column=1).value == i
        assert ws.cell(row=row, column=2).value == line.scope_item.code
        assert ws.cell(row=row, column=4).value == pytest.approx(line.scope_item.quantity)
        assert ws.cell(row=row, column=5).value == line.scope_item.unit
        assert ws.cell(row=row, column=TOTAL_COST_COLUMN).value == pytest.approx(line.total_cost)


def test_currency_format_applied(small_bid_xlsx):
    path, _bid = small_bid_xlsx
    ws = openpyxl.load_workbook(path).active
    # First data row, Total Cost column
    cell = ws.cell(row=11, column=TOTAL_COST_COLUMN)
    assert "$" in cell.number_format


def test_subtotal_and_markups_present(small_bid_xlsx):
    path, bid = small_bid_xlsx
    ws = openpyxl.load_workbook(path, data_only=True).active
    for label in ("SUBTOTAL", "OVERHEAD & PROFIT (20%)", "TAX (8.5%)",
                  "BID BOND (1.5%)", "CONTINGENCIES (5%)", "TOTAL BASE BID"):
        assert _find_cell_by_value(ws, label) is not None, f"missing {label!r}"

    # Numeric values match the in-memory bid.
    subtotal_label = _find_cell_by_value(ws, "SUBTOTAL")
    assert ws.cell(row=subtotal_label.row, column=TOTAL_COST_COLUMN).value == pytest.approx(bid.subtotal)
    oh_label = _find_cell_by_value(ws, "OVERHEAD & PROFIT (20%)")
    assert ws.cell(row=oh_label.row, column=TOTAL_COST_COLUMN).value == pytest.approx(bid.overhead)


# ---------------------------------------------------------------------------
# Edge cases
# ---------------------------------------------------------------------------


def test_empty_bid_still_writes_valid_file(tmp_path, caplog):
    bid = price_bid([])  # no line items, zero subtotal/total
    out = tmp_path / "empty.xlsx"
    with caplog.at_level("WARNING"):
        generate_bid_excel(bid, BidHeader(project_name="Empty"), out)
    assert out.exists()
    assert any("no line items" in r.message for r in caplog.records)

    ws = openpyxl.load_workbook(out, data_only=True).active
    label = _find_cell_by_value(ws, "TOTAL BASE BID")
    assert ws.cell(row=label.row, column=TOTAL_COST_COLUMN).value == 0.0


def test_creates_missing_parent_directories(tmp_path):
    bid = price_bid([ScopeItem("WS5", 1.0, "EA")])
    nested = tmp_path / "deep" / "nested" / "dir" / "bid.xlsx"
    generate_bid_excel(bid, BidHeader(), nested)
    assert nested.exists()


def test_non_bid_input_raises():
    with pytest.raises(TypeError, match="bid must be a Bid"):
        generate_bid_excel("not a bid", BidHeader(), "/tmp/x.xlsx")  # type: ignore[arg-type]


def test_non_header_input_raises(tmp_path):
    bid = price_bid([])
    with pytest.raises(TypeError, match="header must be a BidHeader"):
        generate_bid_excel(bid, "not a header", tmp_path / "x.xlsx")  # type: ignore[arg-type]


def test_description_falls_back_to_unit_cost_table(tmp_path):
    """ScopeItem without a description should pick up the catalog description."""
    bid = price_bid([ScopeItem("WS1", 100.0, "LF")])  # description=""
    out = tmp_path / "desc.xlsx"
    generate_bid_excel(bid, BidHeader(), out)
    ws = openpyxl.load_workbook(out, data_only=True).active
    # First line item description should be non-empty and contain "Cast Stone"
    desc = ws.cell(row=11, column=3).value
    assert desc
    assert "Cast Stone" in desc


def test_description_falls_back_to_code_when_unknown(tmp_path):
    """Unknown code with no description should display the code as a last resort."""
    line_item_scope = [ScopeItem("UNKNOWN_CODE", 1.0, "EA", description="")]
    # We can't price an unknown code; build a Bid manually for this edge case.
    from bid_engine.pricing import BidLineItem, Bid

    line = BidLineItem(
        scope_item=line_item_scope[0],
        unit_labor=10.0,
        unit_material=0.0,
        total_cost=10.0,
    )
    bid = Bid(
        line_items=(line,),
        subtotal=10.0,
        overhead=2.0,
        tax=0.85,
        bid_bond=0.15,
        contingencies=0.5,
        total=13.5,
    )
    out = tmp_path / "unknown.xlsx"
    generate_bid_excel(bid, BidHeader(), out)
    ws = openpyxl.load_workbook(out, data_only=True).active
    assert ws.cell(row=11, column=3).value == "UNKNOWN_CODE"
