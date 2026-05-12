"""Integration tests for bid_engine.pipeline.

The headline test runs the full pipeline on Park Avenue Elementary School
using cached extractions from tests/extractions/ so no vision API calls
happen during testing. The result is read back with openpyxl and the
TOTAL BASE BID cell is verified against the in-memory Bid.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from bid_engine.bid_generator import BidHeader, TOTAL_COST_COLUMN
from bid_engine.pipeline import PipelineResult, run_pipeline
from bid_engine.pricing import DEFAULT_WAGES, Bid, ScopeItem
from bid_engine.scope_checker import ProjectConfig

TESTS_DIR = Path(__file__).resolve().parent
EXTRACTIONS_DIR = TESTS_DIR / "extractions"
BASELINE_PAGE_5 = TESTS_DIR / "baseline_page5.txt"


# ---------------------------------------------------------------------------
# Cached-extraction fake (no API calls)
# ---------------------------------------------------------------------------


def _cached_extractor(_pdf_path: Path, page: int) -> list[dict]:
    """Read tab-separated extractions from disk; mimic extract_page's shape."""
    path = BASELINE_PAGE_5 if page == 5 else EXTRACTIONS_DIR / f"page{page}.txt"
    items: list[dict] = []
    for line in path.read_text().splitlines():
        parts = line.strip().split("\t")
        if len(parts) != 3:
            continue
        code, qty, unit = parts
        items.append({"code": code, "quantity": float(qty), "unit": unit})
    return items


def _failing_extractor(_pdf_path: Path, page: int) -> list[dict]:
    raise RuntimeError(f"simulated extraction failure on page {page}")


def _empty_extractor(_pdf_path: Path, page: int) -> list[dict]:
    return []


def _mixed_extractor(pdf_path: Path, page: int) -> list[dict]:
    """Page 99 fails, every other page is cached."""
    if page == 99:
        raise RuntimeError("simulated failure on page 99")
    return _cached_extractor(pdf_path, page)


# ---------------------------------------------------------------------------
# Headline test
# ---------------------------------------------------------------------------


def test_park_avenue_full_pipeline_via_cached_extractor(tmp_path):
    fake_pdf = tmp_path / "park_ave.pdf"
    fake_pdf.touch()
    out = tmp_path / "park_ave_bid.xlsx"

    result = run_pipeline(
        fake_pdf,
        page_numbers=[2, 3, 5, 6],
        project_config=ProjectConfig(state="NY", stories=3, wage_rates=DEFAULT_WAGES),
        output_path=out,
        header=BidHeader(
            project_name="PARK AVENUE ELEMENTARY SCHOOL",
            address="10 Park Avenue, Warwick, NY 10990",
            date="2026-05-12",
        ),
        extractor=_cached_extractor,
    )

    # Shape
    assert isinstance(result, PipelineResult)
    assert isinstance(result.bid, Bid)
    assert result.output_path == out
    assert out.exists()

    # Aggregation: 24 unique (code, unit) keys from the four pages
    assert len(result.scope_items) == 24
    assert len(result.bid.line_items) == 24

    # Bid total matches the documented pipeline figure within $1
    assert result.bid.total == pytest.approx(35766.51, abs=1.0)

    # Scope alerts: scaffolding/fencing/shed missing (extracted-only scope),
    # boom_lift fires on 3-story, no prevailing_wage alert (DEFAULT_WAGES match)
    alert_ids = {a.item_id for a in result.alerts}
    assert "scaffolding" in alert_ids
    assert "fencing" in alert_ids
    assert "shed_protection" in alert_ids
    assert "boom_lift" in alert_ids
    assert not any(a.item_id.startswith("prevailing_wage") for a in result.alerts)
    assert len(result.alerts) == 4

    # Excel total cell matches the in-memory bid
    ws = openpyxl.load_workbook(out, data_only=True).active
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "TOTAL BASE BID":
                total_cell = ws.cell(row=cell.row, column=TOTAL_COST_COLUMN)
                assert total_cell.value == pytest.approx(result.bid.total)
                break


# ---------------------------------------------------------------------------
# Smaller integration / error-path tests
# ---------------------------------------------------------------------------


def test_pipeline_without_project_config_still_runs(tmp_path):
    fake_pdf = tmp_path / "p.pdf"
    fake_pdf.touch()
    out = tmp_path / "b.xlsx"
    result = run_pipeline(
        fake_pdf,
        page_numbers=[5],
        output_path=out,
        extractor=_cached_extractor,
    )
    assert out.exists()
    # With no ProjectConfig, only the three required-item checks run.
    ids = {a.item_id for a in result.alerts}
    assert ids == {"scaffolding", "fencing", "shed_protection"}


def test_pipeline_continues_when_one_page_fails(tmp_path, caplog):
    fake_pdf = tmp_path / "p.pdf"
    fake_pdf.touch()
    out = tmp_path / "b.xlsx"
    with caplog.at_level("ERROR"):
        result = run_pipeline(
            fake_pdf,
            page_numbers=[5, 99],  # 99 fails
            output_path=out,
            extractor=_mixed_extractor,
        )
    assert out.exists()
    # Page 5 alone produces 11 unique (code, unit) keys after aggregation
    # (21 raw rows on page 5, but several codes repeat across elevations).
    assert len(result.bid.line_items) == 11
    assert any("page 99 extraction failed" in r.message for r in caplog.records)


def test_pipeline_raises_when_no_pages():
    with pytest.raises(ValueError, match="at least one page"):
        run_pipeline("any.pdf", page_numbers=[], extractor=_cached_extractor)


def test_pipeline_raises_on_invalid_page_number():
    with pytest.raises(ValueError, match=">= 1"):
        run_pipeline("any.pdf", page_numbers=[0, 1], extractor=_cached_extractor)


def test_pipeline_raises_when_all_pages_empty(tmp_path):
    fake_pdf = tmp_path / "p.pdf"
    fake_pdf.touch()
    with pytest.raises(RuntimeError, match="no scope items"):
        run_pipeline(
            fake_pdf,
            page_numbers=[1, 2],
            output_path=tmp_path / "b.xlsx",
            extractor=_empty_extractor,
        )


def test_pipeline_raises_when_all_pages_fail(tmp_path):
    fake_pdf = tmp_path / "p.pdf"
    fake_pdf.touch()
    with pytest.raises(RuntimeError, match="no scope items"):
        run_pipeline(
            fake_pdf,
            page_numbers=[1, 2],
            output_path=tmp_path / "b.xlsx",
            extractor=_failing_extractor,
        )


def test_pipeline_missing_pdf_raises_for_real_extractor():
    """With no extractor override, the function should refuse a missing PDF."""
    with pytest.raises(FileNotFoundError, match="PDF not found"):
        run_pipeline(
            "/nonexistent/path.pdf",
            page_numbers=[1],
        )


def test_pipeline_result_immutable(tmp_path):
    fake_pdf = tmp_path / "p.pdf"
    fake_pdf.touch()
    result = run_pipeline(
        fake_pdf,
        page_numbers=[5],
        output_path=tmp_path / "b.xlsx",
        extractor=_cached_extractor,
    )
    with pytest.raises((AttributeError, TypeError)):
        result.bid = None  # type: ignore[misc]
