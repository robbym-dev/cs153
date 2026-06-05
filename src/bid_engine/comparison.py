"""Load a reference bid spreadsheet for side-by-side comparison.

Extends the column convention established in
scripts/compare_extraction_to_spreadsheet.py (load_spreadsheet). That helper
only summed quantity per (code, unit); the web UI also needs total cost,
so we pull QTY (col 6) and TOTAL COST (col 15) per row, summed by
(code, normalized_unit) keyed off the CODE: prefix in the DESCRIPTION cell.
"""

from __future__ import annotations

import logging
import re
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

import openpyxl

from bid_engine.pricing import normalize_unit

logger = logging.getLogger(__name__)

CODE_RE = re.compile(r"^([A-Z]+\d+)\s*:")

DETAIL_SHEET = "DETAIL"
DATA_START_ROW = 28      # first row after the divider row 27
COL_DESC = 5
COL_QTY = 6
COL_UNIT = 7
COL_TOTAL_COST = 15


@dataclass(frozen=True)
class ReferenceLine:
    code: str
    unit: str
    qty: float
    total: float


def load_reference_bid(path: Path) -> dict[tuple[str, str], ReferenceLine]:
    """Sum (qty, total) by (code, normalized_unit) across all coded rows.

    Raises ValueError if the workbook lacks the DETAIL sheet.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    if DETAIL_SHEET not in wb.sheetnames:
        raise ValueError(
            f"reference spreadsheet missing '{DETAIL_SHEET}' sheet "
            f"(found: {wb.sheetnames!r})"
        )
    ws = wb[DETAIL_SHEET]

    qty_totals: dict[tuple[str, str], float] = defaultdict(float)
    cost_totals: dict[tuple[str, str], float] = defaultdict(float)

    for r in range(DATA_START_ROW, ws.max_row + 1):
        desc = ws.cell(r, COL_DESC).value
        qty = ws.cell(r, COL_QTY).value
        unit = ws.cell(r, COL_UNIT).value
        total = ws.cell(r, COL_TOTAL_COST).value
        if not desc or qty is None or not unit:
            continue
        m = CODE_RE.match(str(desc).strip())
        if not m:
            continue
        try:
            qty_f = float(qty)
        except (TypeError, ValueError):
            continue
        try:
            total_f = float(total) if total is not None else 0.0
        except (TypeError, ValueError):
            total_f = 0.0
        key = (m.group(1), normalize_unit(str(unit)))
        qty_totals[key] += qty_f
        cost_totals[key] += total_f

    logger.info(
        "loaded %d (code, unit) line items from reference spreadsheet %s",
        len(qty_totals), path.name,
    )
    return {
        k: ReferenceLine(k[0], k[1], qty_totals[k], cost_totals[k])
        for k in qty_totals
    }
