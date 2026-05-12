"""Bid generator — write a Bid object to an Excel file in Tyler's format.

Layout:
    Row 1   : GENERAL SUMMARY (title)
    Rows 3-7: Project metadata + TOTAL BASE BID
    Row 10  : Column headers (Item #, Code, Description, ...)
    Rows 11+: One per BidLineItem
    Then    : SUBTOTAL, OVERHEAD & PROFIT, TAX, BID BOND, CONTINGENCIES,
              TOTAL BASE BID

Descriptions are taken from `BidLineItem.scope_item.description` when set,
otherwise fall back to the corresponding `UnitCost.description` from
`DEFAULT_UNIT_COSTS`, otherwise the bare code.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from bid_engine.pricing import (
    DEFAULT_UNIT_COSTS,
    Bid,
    BidLineItem,
    normalize_unit,
    waste_rate_for,
)

logger = logging.getLogger(__name__)

CURRENCY_FORMAT = '"$"#,##0.00'
PERCENT_FORMAT = "0.0%"
QTY_FORMAT = "0.00"

# (label, column width in approximate character units). Order is authoritative
# for both the header row and the per-line column ordering below.
COLUMN_SPECS: tuple[tuple[str, int], ...] = (
    ("Item #", 7),
    ("Code", 8),
    ("Description", 60),
    ("Qty", 10),
    ("Unit", 8),
    ("Waste %", 9),
    ("Qty w/ Waste", 13),
    ("Unit Labor", 12),
    ("Unit Material", 14),
    ("Total Labor", 13),
    ("Total Material", 15),
    ("Unit Cost", 11),
    ("Total Cost", 13),
)

TOTAL_COST_COLUMN = len(COLUMN_SPECS)  # 1-indexed: last column

HEADER_ROW = 10
FIRST_LINE_ROW = HEADER_ROW + 1

_BOLD = Font(bold=True)
_BOLD_BIG = Font(bold=True, size=14)
_CENTER = Alignment(horizontal="center")


@dataclass(frozen=True)
class BidHeader:
    project_name: str = "BID PROJECT"
    address: str = ""
    date: str = ""  # free-form; YYYY-MM-DD recommended
    scope: str = "COMPLETE"


def _line_description(line: BidLineItem) -> str:
    si = line.scope_item
    if si.description:
        return si.description
    uc = DEFAULT_UNIT_COSTS.get((si.code, normalize_unit(si.unit)))
    if uc is not None and uc.description:
        return uc.description
    return si.code


def _write_currency(ws, row: int, col: int, value: float, *, bold: bool = False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = CURRENCY_FORMAT
    if bold:
        cell.font = _BOLD
    return cell


def _write_header(ws, header: BidHeader, total: float) -> None:
    ws["A1"] = "GENERAL SUMMARY"
    ws["A1"].font = _BOLD_BIG

    for row, label, value in (
        (3, "PROJECT:", header.project_name),
        (4, "ADDRESS:", header.address),
        (5, "DATE:", header.date),
        (6, "SCOPE:", header.scope),
    ):
        ws.cell(row=row, column=1, value=label).font = _BOLD
        ws.cell(row=row, column=2, value=value)

    ws.cell(row=7, column=1, value="TOTAL BASE BID:").font = _BOLD
    total_cell = ws.cell(row=7, column=2, value=total)
    total_cell.number_format = CURRENCY_FORMAT
    total_cell.font = _BOLD


def _write_column_headers(ws) -> None:
    for col_idx, (label, width) in enumerate(COLUMN_SPECS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=label)
        cell.font = _BOLD
        cell.alignment = _CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_line_items(ws, bid: Bid) -> int:
    """Write all line items; return the next free row."""
    row = FIRST_LINE_ROW
    for i, line in enumerate(bid.line_items, start=1):
        si = line.scope_item
        waste = waste_rate_for(si.unit)
        qty_with_waste = si.quantity * (1 + waste)
        total_labor = qty_with_waste * line.unit_labor
        total_material = qty_with_waste * line.unit_material
        unit_cost = line.unit_labor + line.unit_material

        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=si.code)
        ws.cell(row=row, column=3, value=_line_description(line))
        c = ws.cell(row=row, column=4, value=si.quantity)
        c.number_format = QTY_FORMAT
        ws.cell(row=row, column=5, value=si.unit)
        c = ws.cell(row=row, column=6, value=waste)
        c.number_format = PERCENT_FORMAT
        c = ws.cell(row=row, column=7, value=qty_with_waste)
        c.number_format = QTY_FORMAT
        _write_currency(ws, row, 8, line.unit_labor)
        _write_currency(ws, row, 9, line.unit_material)
        _write_currency(ws, row, 10, total_labor)
        _write_currency(ws, row, 11, total_material)
        _write_currency(ws, row, 12, unit_cost)
        _write_currency(ws, row, 13, line.total_cost)
        row += 1
    return row


def _write_totals(ws, bid: Bid, start_row: int) -> None:
    """Subtotal → markups → TOTAL BASE BID. Labels in col 3, values in last col."""
    row = start_row
    ws.cell(row=row, column=3, value="SUBTOTAL").font = _BOLD
    _write_currency(ws, row, TOTAL_COST_COLUMN, bid.subtotal, bold=True)
    row += 2  # blank gap

    for label, value in (
        ("OVERHEAD & PROFIT (20%)", bid.overhead),
        ("TAX (8.5%)", bid.tax),
        ("BID BOND (1.5%)", bid.bid_bond),
        ("CONTINGENCIES (5%)", bid.contingencies),
    ):
        ws.cell(row=row, column=3, value=label).font = _BOLD
        _write_currency(ws, row, TOTAL_COST_COLUMN, value)
        row += 1

    row += 1  # blank gap before grand total
    ws.cell(row=row, column=3, value="TOTAL BASE BID").font = _BOLD_BIG
    cell = ws.cell(row=row, column=TOTAL_COST_COLUMN, value=bid.total)
    cell.number_format = CURRENCY_FORMAT
    cell.font = _BOLD_BIG


def generate_bid_excel(
    bid: Bid,
    header: BidHeader,
    output_path: Path | str,
) -> Path:
    """Write `bid` to an .xlsx file at `output_path`. Returns the final path."""
    if not isinstance(bid, Bid):
        raise TypeError(f"bid must be a Bid instance, got {type(bid).__name__}")
    if not isinstance(header, BidHeader):
        raise TypeError(
            f"header must be a BidHeader instance, got {type(header).__name__}"
        )

    output_path = Path(output_path)
    if not bid.line_items:
        logger.warning("generate_bid_excel: bid has no line items")

    try:
        output_path.parent.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        raise OSError(
            f"could not create parent directory for {output_path}: {exc}"
        ) from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "BID"

    _write_header(ws, header, bid.total)
    _write_column_headers(ws)
    next_row = _write_line_items(ws, bid)
    _write_totals(ws, bid, start_row=next_row)

    try:
        wb.save(output_path)
    except OSError as exc:
        raise OSError(f"could not write Excel to {output_path}: {exc}") from exc

    logger.info(
        "wrote %d line items + markups to %s (total $%.2f)",
        len(bid.line_items),
        output_path,
        bid.total,
    )
    return output_path
